from RegonAPI import RegonAPI
from openpyxl import load_workbook

api = RegonAPI(bir_version="bir1.1", is_production=True)
try:
    api.authenticate(key="a7b3f12623e54a4bbddd")
except Exception as e:
    raise

wb = load_workbook("Baza_NIP.xlsx")
sheet1 = wb["NIP"]
sheet2 = wb["Dane"]


def search_nip(nip):
    return api.searchData(nip=nip)


def save_data(raw_data: dict):
    new_row_location = sheet2.max_row + 1
    column = 1
    current_row = new_row_location
    if sheet2.cell(column=1, row=1).value is None:
        for key in raw_data.keys():
            sheet2.cell(column=column, row=1, value=key)
            column += 1
        column = 1
    for key in raw_data.keys():
        sheet2.cell(column=column, row=current_row, value=raw_data[key])
        column += 1
    wb.save(filename="Baza_NIP.xlsx")


for row in sheet1.iter_rows(min_row=2, min_col=1, max_col=1):
    for cell in row:
        data = (search_nip(str(cell.value)))
        save_data(data[0])
